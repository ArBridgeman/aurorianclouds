"""Export DataFrame to Excel with optional dropdown (data validation)."""

from pathlib import Path
from typing import Any, BinaryIO, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


def _list_formula(values: list[str]) -> str:
    """Build Excel list formula for data validation (comma-sep, quoted)."""
    # Excel expects formula like "joint,alex,ariel" (commas in values → quoted)
    escaped = [
        str(v).replace('"', '""')
        for v in values
        if v is not None and str(v).strip() != ""
    ]
    return '"' + ",".join(escaped) + '"'


def write_excel_with_validation(
    df: pd.DataFrame,
    path_or_stream: Union[Path, str, BinaryIO],
    validation_config: dict[str, list[Any]] | None = None,
) -> None:
    """
    Write DataFrame to Excel (file or stream) with dropdown validation.

    Args:
        df: DataFrame to export (column order preserved).
        path_or_stream: Output path (.xlsx) or binary stream (e.g. io.BytesIO).
        validation_config: Column name -> list of allowed values. Missing or
            empty list = no validation for that column.
    """
    if not isinstance(path_or_stream, (Path, str)):
        # BinaryIO (e.g. BytesIO)
        target = path_or_stream
    else:
        path = Path(path_or_stream)
        path.parent.mkdir(parents=True, exist_ok=True)
        target = path

    wb = Workbook()
    ws = wb.active
    if ws.title == "Sheet":
        ws.title = "Transactions"

    # Write header and data from DataFrame
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=1
    ):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if validation_config:
        # Column name -> 1-based column index
        col_index: dict[str, int] = {
            name: i for i, name in enumerate(df.columns, start=1)
        }
        n_rows = len(df) + 1  # header + data

        for col_name, allowed in validation_config.items():
            if not allowed or col_name not in col_index:
                continue
            try:
                values = [str(v).strip() for v in allowed if v is not None]
            except (TypeError, ValueError):
                continue
            if not values:
                continue

            formula = _list_formula(values)
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            col_letter = ws.cell(
                row=1, column=col_index[col_name]
            ).column_letter
            dv.add(f"{col_letter}2:{col_letter}{n_rows}")

    wb.save(target)
