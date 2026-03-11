"""Unit tests for banking_helpers.excel_export module."""

import io
from pathlib import Path

import pandas as pd
import pytest
from banking_helpers.excel_export import write_excel_with_validation
from openpyxl import load_workbook


@pytest.fixture
def sample_dataframe():
    """Fixture for sample DataFrame."""
    return pd.DataFrame(
        {
            "Date": ["2026-01-01", "2026-01-02"],
            "Category": ["groceries", "fun"],
            "Payment": ["joint", "joint"],
            "Amount": [-50.00, -10.00],
            "Description": ["rewe markt", "spotify"],
        }
    )


@pytest.fixture
def sample_validation_config():
    """Fixture for sample validation configuration."""
    return {
        "Category": ["groceries", "fun", "household"],
        "Payment": ["joint", "personal"],
    }


class TestWriteExcelBasic:
    """Tests for basic Excel writing functionality."""

    def test_write_to_bytes_stream(self, sample_dataframe):
        """Test writing DataFrame to BytesIO stream."""
        output = io.BytesIO()
        write_excel_with_validation(sample_dataframe, output)
        assert output.tell() > 0
        assert output.getvalue()[:2] == b"PK"  # ZIP file header

    @pytest.mark.parametrize("use_string_path", [False, True])
    def test_write_to_paths(self, sample_dataframe, tmp_path, use_string_path):
        """Test writing DataFrame to file paths."""
        output_path = tmp_path / "test.xlsx"
        if use_string_path:
            output_path = str(output_path)
        write_excel_with_validation(sample_dataframe, output_path)
        assert Path(str(output_path)).exists()

    def test_creates_parent_directories(self, sample_dataframe, tmp_path):
        """Test that parent directories are created."""
        output_path = tmp_path / "subdir" / "nested" / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path)
        assert output_path.exists()


class TestExcelContent:
    """Tests for Excel content correctness."""

    def test_excel_contains_header(self, sample_dataframe, tmp_path):
        """Test that Excel contains column headers."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == list(sample_dataframe.columns)

    def test_excel_contains_data(self, sample_dataframe, tmp_path):
        """Test that Excel contains data rows."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == len(sample_dataframe) + 1  # data + header

    def test_excel_sheet_name_is_transactions(self, sample_dataframe, tmp_path):
        """Test that sheet is named 'Transactions'."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path)

        wb = load_workbook(output_path)
        assert wb.active.title == "Transactions"

    def test_excel_data_matches_dataframe(self, sample_dataframe, tmp_path):
        """Test that Excel data matches DataFrame."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check first data row
        row_values = [cell.value for cell in ws[2]]
        expected = [
            sample_dataframe.iloc[0, i]
            for i in range(len(sample_dataframe.columns))
        ]
        assert row_values == expected


class TestExcelValidation:
    """Tests for data validation (dropdown) functionality."""

    def test_validation_applied_to_column(
        self, sample_dataframe, sample_validation_config, tmp_path
    ):
        """Test that validation is applied to specified column."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, sample_validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        # Check that data validation exists on Category column
        assert len(ws.data_validations.dataValidation) > 0

    def test_validation_formula_contains_values(
        self, sample_dataframe, sample_validation_config, tmp_path
    ):
        """Test that validation formula contains expected values."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, sample_validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        # Find Category column validation
        validations = ws.data_validations.dataValidation
        assert len(validations) > 0

        # Check that formula contains the values
        for validation in validations:
            if validation.formula1:
                assert "groceries" in validation.formula1
                break
        else:
            pytest.fail("No validation with formula found")

    def test_validation_on_multiple_columns(
        self, sample_dataframe, sample_validation_config, tmp_path
    ):
        """Test that validation is applied to every configured column."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, sample_validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        # Collect all cell ranges that have validation
        validated_ranges = " ".join(
            str(dv.sqref) for dv in ws.data_validations.dataValidation
        )

        # Both Category (col B) and Payment (col C) should have dropdowns
        assert "B" in validated_ranges, "Category column missing validation"
        assert "C" in validated_ranges, "Payment column missing validation"

    def test_no_validation_for_empty_list(self, sample_dataframe, tmp_path):
        """Test that no validation is applied when the allowed list is empty."""
        validation_config = {"Category": []}
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        assert len(ws.data_validations.dataValidation) == 0

    def test_validation_ignores_missing_columns(
        self, sample_dataframe, tmp_path
    ):
        """Test that validation config for missing columns is ignored."""
        validation_config = {
            "Category": ["groceries"],
            "NonExistentColumn": ["value"],
        }
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        # Should not crash
        assert ws.max_row > 0

    def test_validation_with_special_characters(
        self, sample_dataframe, tmp_path
    ):
        """Test that validation values with special characters are correctly encoded in the formula."""
        validation_config = {
            "Category": [
                "clothing - special",
                "household - special",
                'value"with"quotes',
            ],
        }
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, validation_config
        )

        wb = load_workbook(output_path)
        ws = wb.active

        validations = ws.data_validations.dataValidation
        assert len(validations) == 1
        formula = validations[0].formula1
        assert "clothing - special" in formula
        assert "household - special" in formula
        # Quotes in values are escaped as ""
        assert 'value""with""quotes' in formula


class TestExcelValidationNone:
    """Tests for handling None validation config."""

    def test_none_validation_config(self, sample_dataframe, tmp_path):
        """Test that None validation config doesn't crash."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path, None)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == len(sample_dataframe) + 1

    def test_empty_validation_config(self, sample_dataframe, tmp_path):
        """Test that empty validation config doesn't crash."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(sample_dataframe, output_path, {})

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == len(sample_dataframe) + 1


class TestExcelEdgeCases:
    """Tests for edge cases and error handling."""

    @pytest.mark.parametrize(
        "df,expected_rows",
        [
            (
                pd.DataFrame(columns=["Date", "Amount"]),
                1,
            ),  # Empty - header only
            (
                pd.DataFrame({"Date": ["2026-01-01"], "Amount": [-50.00]}),
                2,
            ),  # Single row
            (
                pd.DataFrame(
                    {"Date": ["2026-01-01", None], "Amount": [-50.00, -10.00]}
                ),
                3,
            ),  # None values
        ],
    )
    def test_dataframe_variations(self, df, expected_rows, tmp_path):
        """Test handling of various DataFrame structures."""
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == expected_rows

    def test_dataframe_with_various_types(self, tmp_path):
        """Test DataFrame with various data types."""
        df = pd.DataFrame(
            {
                "String": ["test"],
                "Int": [42],
                "Float": [3.14],
                "Bool": [True],
            }
        )
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == 2

    def test_validation_with_none_values_in_list(
        self, sample_dataframe, tmp_path
    ):
        """Test that None values in validation list are filtered."""
        validation_config = {"Category": ["groceries", None, "fun"]}
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(
            sample_dataframe, output_path, validation_config
        )
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row > 0

    def test_large_number_of_rows(self, tmp_path):
        """Test with large number of rows."""
        df = pd.DataFrame(
            {
                "Date": [f"2026-01-{(i % 30) + 1:02d}" for i in range(1000)],
                "Amount": [-50.00] * 1000,
            }
        )
        output_path = tmp_path / "test.xlsx"
        write_excel_with_validation(df, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == 1001  # 1000 rows + header
